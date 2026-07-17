"""
Construction du panel IRIS (groupe de contrôle candidat) à partir des
fichiers INSEE bruts. Contrairement aux QPV, pas de fichier socio séparé :
taux de pauvreté et médiane sont dans le même fichier "revenu disponible".

Entrée : data/raw/iris_{annee}/*.xls(x)
Sortie : data/processed/panel_iris.csv
"""
import xlrd
import openpyxl
import pandas as pd
import glob


def lire_iris_ancien(annee, dossier='data/raw'):
    """Format .xls (2012-2019 environ). En-tête technique toujours en ligne 5,
    mais VÉRIFIER avant de généraliser à de nouvelles années non testées."""
    f = glob.glob(f'{dossier}/iris_{annee}/*{annee}*')[0]
    wb = xlrd.open_workbook(f)
    sh = wb.sheet_by_name('IRIS_DISP')
    header = sh.row_values(5)
    data = [sh.row_values(r) for r in range(6, sh.nrows)]
    df = pd.DataFrame(data, columns=header)
    df = df[df['IRIS'].astype(str).str.len() > 5]
    yy = str(annee)[2:]
    df_slim = df[['IRIS', 'COM', 'LIBCOM', f'DISP_TP60{yy}', f'DISP_MED{yy}']].copy()
    df_slim.columns = ['iris', 'com', 'libcom', 'taux_pauvrete', 'revenu_median']
    df_slim['annee'] = annee
    df_slim['com'] = df_slim['com'].astype(str)
    return df_slim


def lire_iris_recent(annee, dossier='data/raw'):
    """Format .xlsx récent (2021+), structure openpyxl."""
    f = glob.glob(f'{dossier}/iris_{annee}/*.xlsx')[0]
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb['IRIS_DISP']
    data = list(ws.iter_rows(min_row=6, values_only=True))
    df = pd.DataFrame(data[1:], columns=data[0])
    df = df[df['IRIS'].notna()].copy()
    yy = str(annee)[2:]
    df_slim = df[['IRIS', 'COM', 'LIBCOM', f'DISP_TP60{yy}', f'DISP_MED{yy}']].copy()
    df_slim.columns = ['iris', 'com', 'libcom', 'taux_pauvrete', 'revenu_median']
    df_slim['annee'] = annee
    df_slim['com'] = df_slim['com'].astype(str)
    return df_slim


def lire_iris_2016_2019(annee, dossier='data/raw'):
    """Format 2016 (.xls) / 2017-2019 (.xlsx). Détection automatique de la
    ligne d'en-tête technique : la 1ère cellule vaut "IRIS" à la fois sur
    la ligne technique ET sur une ligne "lisible" au-dessus — il faut donc
    aussi vérifier que la 2e colonne vaut "LIBIRIS" pour ne pas se tromper
    de ligne (bug identifié et corrigé, voir carnet section 19).

    NOTE : reconstruite à partir du carnet, non testée dans cet environnement
    faute de fichiers bruts (data/raw non versionné) — à vérifier sur les
    fichiers réels avant de lui faire confiance (voir aussi la note dans
    01_build_panel_qpv.py). Résultat attendu si correcte : cohérent avec
    panel_iris_2016_2019.csv déjà produit et commité (12285/12345/12395/12481
    IRIS pour 2016/2017/2018/2019).
    """
    yy = str(annee)[2:]
    fichiers_xls = glob.glob(f'{dossier}/iris_{annee}/*{annee}*.xls')
    fichiers_xlsx = glob.glob(f'{dossier}/iris_{annee}/*{annee}*.xlsx')

    if fichiers_xls:
        wb = xlrd.open_workbook(fichiers_xls[0])
        feuille = max(wb.sheet_names(), key=lambda s: wb.sheet_by_name(s).nrows)
        sh = wb.sheet_by_name(feuille)
        lignes_apercu = [sh.row_values(r) for r in range(min(15, sh.nrows))]
    else:
        wb = openpyxl.load_workbook(fichiers_xlsx[0], data_only=True)
        feuille = max(wb.sheetnames, key=lambda s: wb[s].max_row)
        ws = wb[feuille]
        lignes_apercu = [list(row) for row in ws.iter_rows(min_row=1, max_row=15, values_only=True)]

    ligne_header = None
    for i, row in enumerate(lignes_apercu):
        c0 = str(row[0]).strip().upper() if row[0] is not None else ''
        c1 = str(row[1]).strip().upper() if len(row) > 1 and row[1] is not None else ''
        if c0 == 'IRIS' and c1 == 'LIBIRIS':
            ligne_header = i
            break
    if ligne_header is None:
        raise ValueError(f"En-tête IRIS/LIBIRIS non trouvé pour {annee} (feuille {feuille})")

    if fichiers_xls:
        header = sh.row_values(ligne_header)
        data = [sh.row_values(r) for r in range(ligne_header + 1, sh.nrows)]
    else:
        header = lignes_apercu[ligne_header]
        data = list(ws.iter_rows(min_row=ligne_header + 2, values_only=True))

    df = pd.DataFrame(data, columns=header)
    df = df[df['IRIS'].notna()].copy()

    tp60_col = cherche_colonne(df.columns, ['disp_tp60_a{yy}', 'disp_tp60{yy}'], yy)
    med_col = cherche_colonne(df.columns, ['disp_med_a{yy}', 'disp_med{yy}'], yy)
    if tp60_col is None or med_col is None:
        raise ValueError(f"Colonnes taux de pauvreté/médiane introuvables pour IRIS {annee}")

    df_slim = df[['IRIS', 'COM', 'LIBCOM', tp60_col, med_col]].copy()
    df_slim.columns = ['iris', 'com', 'libcom', 'taux_pauvrete', 'revenu_median']
    df_slim['annee'] = annee
    df_slim['com'] = df_slim['com'].astype(str)
    return df_slim


def cherche_colonne(colonnes, motifs, yy):
    """Voir 01_build_panel_qpv.py — même logique de résolution de colonne
    tolérante à la casse et à la convention de nommage du millésime."""
    cibles = {m.format(yy=yy).lower() for m in motifs}
    for c in colonnes:
        if str(c).lower() in cibles:
            return c
    return None


if __name__ == '__main__':
    panels = []
    for annee in [2012, 2013, 2014]:
        panels.append(lire_iris_ancien(annee))
    for annee in [2016, 2017, 2018, 2019]:
        panels.append(lire_iris_2016_2019(annee))
    for annee in [2021]:
        panels.append(lire_iris_recent(annee))

    panel_complet = pd.concat(panels, ignore_index=True)
    panel_complet.to_csv('data/processed/panel_iris_complet.csv', index=False)
    print(f"Panel IRIS construit : {len(panel_complet)} lignes")
    print(panel_complet.groupby('annee').size())
