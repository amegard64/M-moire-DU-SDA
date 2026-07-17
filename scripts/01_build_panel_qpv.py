"""
Construction du panel QPV (traité) à partir des fichiers INSEE bruts.
Gère les incohérences de structure entre millésimes (position d'en-tête,
noms de colonnes, présence variable de Gini/S80-S20).

Entrée attendue : fichiers .xls/.xlsx bruts INSEE, un dossier par année,
sous data/raw/qpv_{annee}/
Sortie : data/processed/panel_qpv_2012_2021.csv (ou sous-ensemble selon
les années disponibles)
"""
import xlrd
import openpyxl
import pandas as pd
import glob

DOM_PREFIXES = ('QP971', 'QP972', 'QP973', 'QP974', 'QP975',
                'QP976', 'QP977', 'QP978', 'QP987', 'QP988')


def lire_xls_intelligent(fichier, feuille=None):
    """Lit un fichier .xls ancien format en détectant automatiquement la
    ligne d'en-tête technique (celle dont la 1ère cellule vaut QP/CODGEO/IRIS).
    Ne JAMAIS fixer un numéro de ligne en dur : ça varie selon le millésime.
    """
    wb = xlrd.open_workbook(fichier)
    if feuille is None:
        feuille = [s for s in wb.sheet_names() if wb.sheet_by_name(s).nrows > 100][0]
    sh = wb.sheet_by_name(feuille)
    ligne_header = None
    for i in range(min(15, sh.nrows)):
        val = str(sh.cell_value(i, 0)).strip().upper()
        if val in ('QP', 'CODGEO', 'IRIS'):
            ligne_header = i
            break
    if ligne_header is None:
        raise ValueError(f"En-tête non trouvé dans {fichier} (feuille {feuille})")
    header = [str(c) for c in sh.row_values(ligne_header)]
    data = [sh.row_values(r) for r in range(ligne_header + 1, sh.nrows)]
    df = pd.DataFrame(data, columns=header)
    code_col = header[0]
    df['code'] = df[code_col].astype(str).str.upper()
    df = df[df['code'].str.startswith(('QP', 'IRIS'))].copy()
    return df


def lire_xlsx_2021(fichier, feuille='Données quartiers'):
    """Lit le format xlsx récent (2021+), structure différente des .xls anciens."""
    wb = openpyxl.load_workbook(fichier, data_only=True)
    ws = wb[feuille]
    data = list(ws.iter_rows(min_row=6, values_only=True))
    header = data[0]
    rows = data[1:]
    df = pd.DataFrame(rows, columns=header)
    df = df[df['CODGEO'].notna()].copy()
    df['code'] = df['CODGEO'].astype(str).str.upper()
    return df


def lire_xlsx_intelligent(fichier):
    """Lit un fichier .xlsx (2017-2019) en détectant automatiquement la feuille
    et la ligne d'en-tête technique, sans supposer un nom de feuille ou une
    position de ligne stables (voir carnet section 14 : noms de feuille et
    positions d'en-tête varient d'un millésime à l'autre)."""
    wb = openpyxl.load_workbook(fichier, data_only=True)
    feuille = max(wb.sheetnames, key=lambda s: wb[s].max_row)
    ws = wb[feuille]
    ligne_header = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True)):
        val = str(row[0]).strip().upper() if row[0] is not None else ''
        if val in ('QP', 'CODGEO', 'IRIS'):
            ligne_header = i + 1
            break
    if ligne_header is None:
        raise ValueError(f"En-tête non trouvé dans {fichier} (feuille {feuille})")
    data = list(ws.iter_rows(min_row=ligne_header, values_only=True))
    df = pd.DataFrame(data[1:], columns=data[0])
    code_col = data[0][0]
    df = df[df[code_col].notna()].copy()
    df['code'] = df[code_col].astype(str).str.upper()
    return df


def cherche_colonne(colonnes, motifs, yy):
    """Cherche une colonne parmi plusieurs motifs possibles (insensible à la
    casse), car le nom technique change de convention selon les millésimes
    (ex. DISP_Q2_A17 devient DISP_MED_A18 - la médiane change de nom, pas
    seulement de casse). Échoue explicitement plutôt que de deviner en
    silence si aucun motif ne correspond (voir carnet : "ne jamais supposer
    une correspondance directe de nom de colonne entre deux millésimes")."""
    cibles = {m.format(yy=yy).lower() for m in motifs}
    for c in colonnes:
        if c.lower() in cibles:
            return c
    return None


def extraire_variables_qpv(annee, dossier_donnees='data/raw'):
    """Extrait revenu_median, taux_pauvrete, gini, s80s20 pour une année donnée.
    Gère les changements de nom de colonne et de fichier source selon le millésime :
    - 2012-2014 : médiane/gini dans le fichier 'revenu disponible', taux de
      pauvreté dans le fichier 'socio' séparé
    - 2016-2019 : taux de pauvreté déjà dans le fichier 'revenu disponible'
      (bascule entre 2014 et 2016, voir carnet section 19) ; 2016 en .xls,
      2017-2019 en .xlsx ; nom ET casse de la colonne médiane changent
      entre 2017 (disp_q2_a17) et 2018 (DISP_MED_A18)
    - 2021+ : tout dans le fichier 'revenu disponible', colonnes en majuscules

    NOTE : la branche 2016-2019 ci-dessous a été reconstruite à partir des
    observations consignées dans le carnet (section 19), faute de disposer
    des fichiers bruts dans cet environnement pour la tester directement
    (data/raw n'est pas versionné). À vérifier en la faisant tourner sur les
    fichiers réels avant de lui faire confiance : le résultat attendu est
    1296 QPV pour chacune des années 2016 à 2019 (cf. panel_qpv_2016_2019.csv
    déjà produit et commité).
    """
    yy = str(annee)[2:]
    dossier = f"{dossier_donnees}/qpv_{annee}"

    if annee <= 2014:
        f_rev = glob.glob(f'{dossier}/*revenu-disponible*{annee}*')[0] if annee != 2013 \
            else glob.glob(f'{dossier}/*disponible*{annee}*')[0]
        f_soc = glob.glob(f'{dossier}/*socio*{annee}*')[0]
        rev = lire_xls_intelligent(f_rev)
        soc = lire_xls_intelligent(f_soc)

        med_col = next((c for c in rev.columns if c.lower() == f'disp_q2_a{yy}'), None)
        gini_col = next((c for c in rev.columns if 'gini' in c.lower()), None)
        s80s20_col = next((c for c in rev.columns if 's80s20' in c.lower()), None)
        tp60_col = next((c for c in soc.columns if c.lower() == f'tp60_a{yy}'), None)

        rev_slim = rev[['code'] + [c for c in [med_col, gini_col, s80s20_col] if c]].copy()
        rev_slim.columns = ['code'] + [n for n, c in
                                        [('revenu_median', med_col), ('gini', gini_col),
                                         ('s80s20', s80s20_col)] if c]
        soc_slim = soc[['code', tp60_col]].copy()
        soc_slim.columns = ['code', 'taux_pauvrete']

        panel = rev_slim.merge(soc_slim, on='code', how='outer')
    elif annee <= 2019:
        motif = f'{dossier}/*disponible*{annee}*.xls'
        f_rev = glob.glob(motif)[0] if glob.glob(motif) else glob.glob(f'{dossier}/*disponible*{annee}*.xlsx')[0]
        rev = lire_xls_intelligent(f_rev) if f_rev.endswith('.xls') else lire_xlsx_intelligent(f_rev)

        med_col = cherche_colonne(rev.columns, ['disp_med_a{yy}', 'disp_q2_a{yy}', 'disp_med{yy}'], yy)
        tp60_col = cherche_colonne(rev.columns, ['disp_tp60_a{yy}', 'disp_tp60{yy}'], yy)
        gini_col = cherche_colonne(rev.columns, ['disp_gi_a{yy}', 'disp_gini_a{yy}'], yy)
        s80s20_col = cherche_colonne(rev.columns, ['disp_s80s20_a{yy}'], yy)
        for nom, col in [('médiane', med_col), ('taux de pauvreté', tp60_col)]:
            if col is None:
                raise ValueError(f"Colonne {nom} introuvable pour {annee} dans {f_rev} "
                                  f"— vérifier la convention de nommage réelle du fichier")

        panel = rev[['code', med_col, tp60_col] + [c for c in [gini_col, s80s20_col] if c]].copy()
        panel.columns = ['code', 'revenu_median', 'taux_pauvrete'] + \
            [n for n, c in [('gini', gini_col), ('s80s20', s80s20_col)] if c]
    else:
        f_rev = glob.glob(f'{dossier}/*disponible*{annee}*.xlsx')[0]
        rev = lire_xlsx_2021(f_rev)
        panel = rev[['code', f'DISP_MED_A{yy}', f'DISP_TP60{yy}',
                     f'DISP_GI_A{yy}', f'DISP_S80S20_A{yy}']].copy()
        panel.columns = ['code', 'revenu_median', 'taux_pauvrete', 'gini', 's80s20']

    panel = panel[~panel['code'].str.startswith(DOM_PREFIXES)].copy()
    panel['annee'] = annee
    return panel


if __name__ == '__main__':
    annees_disponibles = [2012, 2013, 2014, 2016, 2017, 2018, 2019, 2021]  # 2015 et 2020 exclus, voir carnet
    panels = [extraire_variables_qpv(a) for a in annees_disponibles]
    panel_complet = pd.concat(panels, ignore_index=True)
    panel_complet.to_csv('data/processed/panel_qpv_complet.csv', index=False)
    print(f"Panel QPV construit : {len(panel_complet)} lignes, années {annees_disponibles}")
